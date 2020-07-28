"""Contains ExcelFile class for handling MS Excel file.

"""
import re
import logging
import string
from tempfile import TemporaryFile
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
logger = logging.getLogger(__name__)


def int2letters(n):
    """Converts an integer to a string like the MS excel column letters.
    It is essentially a numeral system with a base of 26.
    Examples:
        1 -> A, 2 -> B, ..., 26 -> Z, 27 -> AA, 28 -> AB, ..., 705 -> AAA

    Args:
        n: An integer number (with a base of 10).

    Returns (str): A string like the MS excel column letters

    """
    suffix = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        suffix = chr(r + ord('A')) + suffix
    return suffix


def letters2int(letters):
    """Converts a string of Excel column like letters to integer.
    Examples:
        A -> 1, B -> 2, ..., Z -> 26, AA -> 27, AB -> 28, ..., AAA -> 705

    Args:
        letters: A string contains only ASCII letters.

    Returns: An integer representing the input string with a base of 10.

    """
    n = 0
    for c in letters:
        if c not in string.ascii_letters:
            raise ValueError("Input can only contain ASCII letters. Invalid character: %s." % c)
        n = n * 26 + (ord(c.upper()) - ord('A')) + 1
    return n


class ExcelFile:
    """Represents an MS Excel file.
    This class is built based on openpyxl.

    Attributes:
        file_path: Path of the excel file.
        workbook: openpyxl Workbook object of the excel file.
            See https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/workbook/workbook.html
        worksheet: openpyxl Worksheet object of the active spreadsheet.
            See https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.worksheet.html

    """
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def __init__(self, file_path=None, read_only=False):
        """Initializes an Excel file.
        The workbook will be loaded if the file_path is not None.
        Otherwise a new workbook will be created in the memory.

        Args:
            file_path: The path of the file.
            read_only: Indicates if the file should be opened in read-only mode.
        """
        self.file_path = file_path
        if self.file_path:
            self.workbook = load_workbook(
                filename=self.file_path,
                data_only=True,
                read_only=read_only
            )
        else:
            # Initialize new workbook if file path is not specified.
            self.workbook = Workbook()
        self.worksheet = None
        self.headers = None
        self.set_worksheet()

    def set_worksheet(self, name=None):
        if name:
            self.worksheet = self.workbook[name]
        else:
            self.worksheet = self.workbook.active
        self.headers = self.get_row_values(1)

    def set_headers(self, row_number):
        """Uses a particular row in the file as header row.
        This method modifies the self.headers attribute.

        Args:
            row_number: 1-based row number.

        Returns: Headers as a list of strings.
            An empty list will be returned if the row does not exist.

        """
        self.headers = self.get_row_values(row_number)
        return self.headers

    def save(self, save_as_file_path=None):
        """Saves the workbook.

        Args:
            save_as_file_path (str, Optional): Path of the output file.
            If save_as_file_path is None, changes will be saved into the original file (self.file_path will be used).
            If save_as_file_path is specified:
                1. the contents including the changes, will be saved as another file.
                2. the self.file_path will be changed to save_as_file_path
            If save_as_file_path is None and self.file_path is None, there will be an error.

        """
        if save_as_file_path:
            self.workbook.save(save_as_file_path)
            self.file_path = save_as_file_path
        else:
            if self.file_path:
                self.workbook.save(self.file_path)
            else:
                raise ValueError("File path must be specified.")

    def content(self):
        """The contents of the file, as returned from file.read().

        Returns (bytes): The content of the file. This can be used for HTTP response.

        """
        with TemporaryFile() as temp_file:
            self.workbook.save(temp_file)
            temp_file.seek(0)
            return temp_file.read()

    def column_index(self, header, case_sensitive=False):
        """Gets the 1-based index of a column in the file.

        Args:
            header: The header (name) of the column.
            case_sensitive: Indicates whether the matching should be case sensitive.

        Returns: 1-based index of the column. Or -1 if the header is not found.

        """
        index = -1
        for i, value in enumerate(self.headers):
            if case_sensitive and re.fullmatch(header, value):
                index = i + 1
            elif not case_sensitive and re.fullmatch(header, value, re.IGNORECASE):
                index = i + 1
        return index

    def export_rows(self, row_list):
        """Exports specific rows from the file to a new workbook.

        Args:
            row_list: A list of row numbers.

        Returns: An openpyxl Workbook object.

        Remarks: Only the plain text are exported. Styles and formatting are not copied/exported.

        """
        in_ws = self.worksheet
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.append([cell.value for cell in in_ws[1]])
        for row in row_list:
            if row is None or row <= 0:
                continue
            cells = [cell.value if cell.value is not None else '' for cell in in_ws[int(row)]]
            out_ws.append(cells)
        return out_wb

    def get_row_values(self, row_number):
        """Gets the values of a row as a list of strings.
        Empty cell (None value) will be converted to empty string.
        The leading and trailing whitespaces of the cell value will be removed.

        Args:
            row_number: 1-based row number.

        Returns: A list of strings.
            The returned list will be empty if the row does not exist in the file.

        """
        row = None
        for i, r in enumerate(self.worksheet.rows):
            if row_number <= 0:
                break
            if row_number == i + 1:
                row = r
                break
        if row:
            values = [str(item.value).strip() if item.value is not None else "" for item in row]
        else:
            logger.debug("The excel file does not have row %d" % row_number)
            values = []
        return values

    def has_headers(self, headers, flags=0):
        """Checks if the first row of the active worksheet has certain headers (a list of words).
        A header is considered found if it matches the value of a cell in the first row.
        The matching will be done using re.fullmatch(header, value, flags)
        The leading and trailing whitespaces of the cell value will be removed before matching.

        Args:
            headers: A list of strings, each is a header.
            flags: Flags for python regular expression matching.

        Returns: True if all strings in headers are found in the first row. Otherwise False.

        See Also:
            https://docs.python.org/3/library/re.html#re.fullmatch
            https://docs.python.org/3/library/re.html#re.compile

        """
        for header in headers:
            match = None
            for value in self.headers:
                match = re.fullmatch(header, value, flags)
                if match is not None:
                    break
            if match is None:
                logger.error("Column \"%s\" not found in the Excel file." % header)
                logger.debug("Columns in the file: %s" % ",".join(self.headers))
                return False
        return True

    def get_data_table(self, skip_first_row=True):
        """Reads the excel file into a 2D list in memory.
        This method may not be suitable for large file.

        Args:
            skip_first_row: Indicates whether the first row should be skipped.

        """
        table = []
        empty_count = 0

        for index, row in enumerate(self.worksheet.rows):
            if index == 0 and skip_first_row:
                continue
            values = []
            empty_row = True
            for item in row:
                values.append(item.value)
                if item.value:
                    empty_row = False
            if not empty_row:
                empty_count = 0
            else:
                empty_count += 1
                if empty_count > 200:
                    logger.error(
                        "More than consecutive 200 empty line found in the excel file. "
                        "Data might be truncated."
                    )
                    break
            table.append(values)
        return table

    def write_row(self, value_list, row_number, **kwargs):
        """Writes a list of values to a particular row.
        
        Args:
            value_list (list): A list of values
            row_number (int): 1-based row number
        
        Returns:
            [type]: [description]
        """
        cells = []
        for col, val in enumerate(value_list, start=1):
            cell = self.worksheet.cell(row=row_number, column=col, value=val)
            for attr, value in kwargs.items():
                setattr(cell, attr, value)
            cells.append(cell)
        return cells

    def append_row(self, value_list, **kwargs):
        """Appends a list of values as a row in the file.
        
        Args:
            value_list (list): A list of values to be appended.
        
        Returns:
            list: A list of cells.
        """
        row_number = self.worksheet.max_row + 1
        return self.write_row(value_list, row_number, **kwargs)

    @staticmethod
    def __update_column_width(row, column_widths):
        for i, cell in enumerate(row, start=1):
            cell_size = len(str(cell.value))
            if not cell_size:
                continue
            if cell_size > column_widths.get(i, 0):
                column_widths[i] = cell_size
        return column_widths

    def auto_column_width(self, min_width=10, max_width=100):
        """Automatically sets the columns width base on the number of characters.
        
        Args:
            min_width (int, optional): Minimum width. Defaults to 10.
            max_width (int, optional): Maximum width. Defaults to 100.
        """
        column_widths = {}
        for row in self.worksheet.rows:
            column_widths = self.__update_column_width(row, column_widths)

        for col, column_width in column_widths.items():
            if column_width > max_width:
                column_width = max_width
            if column_width < min_width:
                column_width = min_width
            self.worksheet.column_dimensions[get_column_letter(col)].width = column_width * 1.1
